# Copyright (c) 2022 Shuhei Nitta. All rights reserved.
import os
import re
import enum
import json
import pathlib
import datetime
import concurrent.futures as cf
import typing as t

import click
import openpyxl
import openpyxl.utils
import openpyxl.styles.numbers
import pandas as pd
from yahoo_auction_auto import YahooAuction, cookie, info

__version__ = "0.1.0"


APPNAME = "ヤフオク!在庫表作成アプリ"
APPDIR = pathlib.Path(click.get_app_dir(APPNAME, roaming=False))
LOGDIR = APPDIR / "logs"


@click.command(
    name="",
    help=APPNAME
)
@click.option(
    "-u", "--username",
    type=str,
    prompt="ユーザー名を入力してください",
    help="ヤフオク!のユーザー名"
)
@click.option(
    "-c", "--cookiefile",
    type=click.types.Path(),
    default="",
    help="cookies.jsonのパス"
)
@click.option(
    "--costrate",
    type=float,
    default=0.2,
    help="オークション開始価格に対する仕入れ価格の割合"
)
@click.option(
    "--open-xlsx",
    is_flag=True,
    help="作成されたExcelファイルを開く"
)
@click.version_option(version=__version__, prog_name=APPNAME)
def app(
    username: str,
    cookiefile: str,
    costrate: int,
    open_xlsx: bool
) -> None:
    APPDIR.mkdir(exist_ok=True)
    click.echo("ログイン情報を確認しています...")
    cookiefile = cookiefile or (APPDIR / f"{username}.cookies.json").as_posix()
    ya = get_yahoo_auction(cookiefile)
    aIDs = ya.get_aIDs_selling()
    click.echo(f"出品中の商品: {len(aIDs)}件")
    with cf.ThreadPoolExecutor(8) as executor:
        futures = [executor.submit(ya.get_info_selling, aID) for aID in aIDs]
        with click.progressbar(
            cf.as_completed(futures),
            length=len(futures),
            label="データを取得しています..."
        ) as bar:
            df = pd.concat([parse_info(future.result(), costrate) for future in bar])
    click.echo("ワークブックを作成しています...")
    wb = dataframe_to_workbook(df)
    wb_filename = f"{username}_{datetime.date.today().isoformat()}.xlsx"
    wb.save(wb_filename)
    click.echo(f"{wb_filename}を作成しました")
    if open_xlsx:
        click.launch(wb_filename)


class Column(enum.Enum):
    TITLE = "商品名"
    STOCK = "在庫数"
    COST = "仕入れ価格（円）"
    PRICE = "販売価格（円）"


COLUMN_SIZE = {
    Column.TITLE: 100,
    Column.STOCK: 10,
    Column.PRICE: 16,
    Column.COST: 16
}


def get_yahoo_auction(filename: str | os.PathLike[str]) -> YahooAuction:
    try:
        with open(filename) as f:
            cookies = json.load(f)
        if not YahooAuction(cookies).islogin():
            raise Exception("Invalid cookies")
    except Exception:
        cookies = cookie.get_cookies()
        with open(filename, "w") as f:
            json.dump(cookies, f, indent=2)
    return YahooAuction(cookies)


def parse_info(
    info: info.InfoSelling,
    costrate: float
) -> pd.DataFrame:
    df = pd.DataFrame()
    df[Column.TITLE] = [info.title]
    df[Column.STOCK] = [info.stock]
    if match := re.search(r"[0-9,]+(?= 円)", info.startprice):
        price = int(match.group(0).replace(",", ""))
        df[Column.PRICE] = [price]
        df[Column.COST] = [int(price * costrate)]
    else:
        df[Column.PRICE] = [None]
        df[Column.COST] = [None]
    return df[list(Column)]


def dataframe_to_workbook(df: pd.DataFrame) -> openpyxl.Workbook:
    total_cost = df[Column.COST].fillna(0).dot(df[Column.STOCK])
    total_price = df[Column.PRICE].fillna(0).dot(df[Column.STOCK])

    def total(col: Column) -> t.Any:
        if col is Column.COST:
            return total_cost
        elif col is Column.PRICE:
            return total_price
        else:
            return ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([str(col.value) for col in df.columns])
    ws.append([total(col) for col in df.columns])
    for _, row in df.iterrows():
        ws.append(list(row))
    for idx, col in enumerate(df.columns, 1):
        column_letter = openpyxl.utils.get_column_letter(idx)
        ws.column_dimensions[column_letter].width = COLUMN_SIZE[col]
        if col in [Column.PRICE, Column.COST]:
            for cell in ws[column_letter]:
                cell.number_format = "#,###"
    return wb


if __name__ == "__main__":
    app()
